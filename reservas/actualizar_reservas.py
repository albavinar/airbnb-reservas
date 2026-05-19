#!/usr/bin/env python3
"""
Actualiza RESERVAS_2026.xlsx y LIMPIEZAS_PROXIMAS.xlsx
leyendo emails de Gmail (últimas 48h).
Usa el claude CLI local para parsear emails — no necesita API key.
"""

import argparse, base64, datetime, io, json, os, re, shutil, subprocess, sys, time
from pathlib import Path


def _retry(fn, max_attempts=3, base_delay=2):
    """Reintenta fn() con backoff exponencial. Lanza la última excepción si todo falla."""
    last_exc = None
    for attempt in range(max_attempts):
        try:
            return fn()
        except Exception as exc:
            last_exc = exc
            if attempt < max_attempts - 1:
                time.sleep(base_delay ** attempt)
    raise last_exc

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# ── Paths ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR    = Path(__file__).parent
_default_data = Path.home() / "Documents/Claude/Projects/AIRBNB herramientas"
BASE_DIR      = Path(os.environ.get("DATA_DIR", str(_default_data)))
RESERVAS_XLS  = BASE_DIR / "RESERVAS_2026.xlsx"
LIMPIEZAS_XLS = BASE_DIR / "LIMPIEZAS_2026.xlsx"
CREDS_FILE    = SCRIPT_DIR / "credentials.json"
TOKEN_FILE    = SCRIPT_DIR / "token.json"
_default_log  = Path.home() / "Library/Logs/reservas-airbnb.log" if sys.platform == "darwin" else Path("/tmp/reservas-airbnb.log")
LOG_FILE      = Path(os.environ.get("LOG_FILE", str(_default_log)))

# ── Google OAuth ──────────────────────────────────────────────────────────────

SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/drive",
]

# ── Drive IDs ─────────────────────────────────────────────────────────────────

DRIVE_RESERVAS_PARENT  = "1jXRPN1HjgWSmk2LTmVScyOaMaEtqlGax"  # carpeta BUSSINESS
DRIVE_LIMPIEZAS_PARENT = "1fOZ-WPvEwnJZCioXDWUm32u_ZB_pP1qv"  # carpeta compartida Limpiezas Alba y Gisele
DRIVE_CACHE_PARENT     = "15stGrk0IgfNAGupejzvpjVir0a-XqOT6"   # carpeta bot — sincroniza data_cache.json

# ── Configuración de apartamentos ─────────────────────────────────────────────

SHEETS = ["Sweet 2026", "Sunset 2026", "Center 2026", "Open Sky 2026", "Duplex 2026"]
DRIVE_MAP = {
    "Sweet 2026":    "SWEET APT",
    "Sunset 2026":   "SUNSET APT",
    "Center 2026":   "CENTER B",
    "Open Sky 2026": "OPEN SKY",
    "Duplex 2026":   "DUPLEX",
}
COL_FORMATS = [
    (1,  "dd/mm/yyyy", "center"), (2,  "dd/mm/yyyy", "center"),
    (3,  "General",    "left"),   (4,  "0",           "center"),
    (5,  "#,##0.00",   "center"), (6,  "General",     "center"),
    (7,  "0",          "center"), (8,  "0",            "center"),
    (9,  "#,##0.00",   "center"), (10, "General",      "center"),
    (11, "General",    "left"),
]
GREEN, WHITE, BLUE = "FFD9EAD3", "FFFFFFFF", "FFCFE2F3"

APT_COLORS = {
    "SWEET APT":  "FF48B3FF",
    "SUNSET APT": "FFE06666",
    "CENTER B":   "FFFF9900",
    "OPEN SKY":   "FF00FF00",
    "DUPLEX":     "FFFFFF00",
}
LIMP_GREY  = "FFF6F8F9"
LIMP_HECHA = "FFD9D9D9"  # gris para fila de limpieza ya realizada
DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
# Valores que genera el script — cualquier otro valor se considera edición manual
# Open Sky y Duplex usan columnas distintas (gestionados manualmente en Drive)
OS_DUPLEX_SHEETS = {"Open Sky 2026", "Duplex 2026"}

# Posiciones de adultos/niños/extras por sheet (1-indexed, para refresh_montaje y cache)
SHEET_COL_MAP = {
    "Open Sky 2026": {"adultos": 5, "ninos": 6, "extras": 10},
    "Duplex 2026":   {"adultos": 5, "ninos": 6, "extras": 10},
}
_DEFAULT_COLS = {"adultos": 7, "ninos": 8, "extras": 12}

_AUTO_MONTAJE = {"", "1 CAMA", "2 CAMAS"}


def camas_label(adultos, ninos, extras=""):
    total = (adultos or 0) + (ninos or 0)
    lavanderia = "25" in str(extras or "")
    return "2 CAMAS" if (lavanderia or total >= 3) else "1 CAMA"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── Telegram ─────────────────────────────────────────────────────────────────

def _load_telegram_cfg():
    cfg_path = SCRIPT_DIR / "config.json"
    if not cfg_path.exists():
        return None, None
    try:
        cfg = json.load(cfg_path.open())
        return cfg.get("telegram_token"), cfg.get("chat_id")
    except Exception:
        return None, None

def send_telegram(msg):
    import urllib.request, urllib.parse
    token, chat_id = _load_telegram_cfg()
    if not token or not chat_id:
        return
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    data = urllib.parse.urlencode({"chat_id": chat_id, "text": msg, "parse_mode": "HTML"}).encode()
    try:
        urllib.request.urlopen(url, data=data, timeout=10)
    except Exception:
        pass


# ── Claude CLI ────────────────────────────────────────────────────────────────

def find_claude_bin():
    """Finds the claude CLI binary — checks PATH first, then Application Support."""
    path = shutil.which("claude")
    if path:
        return path
    base = Path.home() / "Library/Application Support/Claude/claude-code"
    if base.exists():
        for v in sorted(base.iterdir(), key=lambda p: p.name, reverse=True):
            candidate = v / "claude.app/Contents/MacOS/claude"
            if candidate.exists():
                return str(candidate)
    return None

CLAUDE_BIN = find_claude_bin()

PARSE_PROMPT_TEMPLATE = """Eres un extractor de datos de reservas de apartamentos turísticos.

Apartamentos: Sweet, Sunset, Center, Open Sky, Duplex.
Plataformas:  Airbnb, Booking, Holidu.
Aliases de apartamentos:
- Sweet   → "Sweet Apartment", "Sweet apt", "a pie de playa 50m", "Sweet apt- Sea Views"
- Sunset  → "Ático deluxe", "vistas al mar, SPA", "BBQ privado", "Sunset Apartment", "Sunset apt"
- Center  → "Center beach", "a una calle de la playa", "Center Apartment", "Center apt"
- Open Sky → "Open Sky Apartment", "Open Sky apt" (ATENCIÓN: Open Sky y Duplex NO reciben confirmaciones de reserva por email — sus reservas se gestionan manualmente. Si el email parece ser una nueva reserva para Open Sky o Duplex, es casi seguro un error de clasificación: revisa el contenido con cuidado. Si menciona BBQ privado, SPA, ático o terraza con vistas al mar → es Sunset. En caso de duda devuelve apartamento: null.)

Reglas para clasificar pagos:
- tipo "nueva_reserva": confirmación de nueva reserva.
- tipo "cancelacion": cancelación de reserva existente. Extrae el importe que el anfitrión conserva
  en el campo cobro_cancelacion (p.ej. "tu cobro se ha actualizado a 87,21 €" → 87.21).
  Si el reembolso es total y no se retiene nada, cobro_cancelacion es 0 o null.
- tipo "modificacion": cambio confirmado de fechas, noches o importe de una reserva ya existente.
  Incluye los datos NUEVOS en los campos correspondientes. Además rellena fecha_entrada_anterior
  con la fecha de entrada original si aparece en el email (para poder localizar la reserva).
- tipo "extra": pagos de servicios adicionales. Importes de extra reconocidos: 25€ (lavandería),
  30€ (check-in anticipado), 100€ (mascota).
  En emails "Has solicitado dinero a X" o "X ha aceptado tu solicitud para enviarte dinero":
    1. Si el importe es exactamente 25€, 30€ o 100€ → extra directo.
    2. Si el importe NO coincide pero al restarle un múltiplo de 3,5€ (tasa turística) el
       resultado es 25€, 30€ o 100€ → es tasa + extra combinados. Clasifica como "extra"
       y pon en extra_importe SOLO la parte del extra (25, 30 o 100), ignorando la tasa.
       Ejemplo: 32€ enviados, tasa probable 7€ (2 personas × 1 noche) → extra_importe=25 (lavandería).
    3. Cualquier otro importe que no encaje con ningún patrón → pon tipo "extra",
       extra_importe=null y ambiguo=true con una nota descriptiva del importe y del huésped
       para que el anfitrión pueda decidir manualmente.
  El extra_canal es siempre "airbnb" cuando el pago llega por notificación de Airbnb.
- tipo "irrelevante": TODO lo demás, incluyendo:
    * Depósitos de seguridad (el depósito de 200€ del Sunset, cualquier fianza)
    * Tasa turística (importes típicos: múltiplos de 3,5€ por persona/noche)
    * Emails "Has solicitado / ha aceptado" con importe distinto de 25€, 30€ o 100€
    * Pagos globales de reserva (el importe total de la estancia)
    * Notificaciones sin pago
    * Cualquier pago cuyo importe coincida con el total de la reserva o con una tasa/depósito conocido
- Bebés se suman a niños.
- Si un pago de Airbnb mezcla tasa turística + posible extra sin desglose claro → ambiguo: true.

Asunto: {subject}
Email:
{body}

Responde SOLO con JSON válido, sin markdown ni explicaciones:
{{
  "tipo": "nueva_reserva" | "cancelacion" | "modificacion" | "extra" | "irrelevante",
  "fecha_entrada_anterior": "DD/MM/YYYY" | null,
  "apartamento": "Sweet" | "Sunset" | "Center" | "Open Sky" | "Duplex" | null,
  "nombre": null,
  "fecha_entrada": "DD/MM/YYYY" | null,
  "fecha_salida": "DD/MM/YYYY" | null,
  "noches": null,
  "total_eur": null,
  "plataforma": "Airbnb" | "Booking" | "Holidu" | "Directa" | null,
  "adultos": null,
  "ninos": null,
  "extra_importe": null,
  "extra_canal": "airbnb" | null,
  "extra_concepto": "lavanderia" | "checkin_anticipado" | "mascota" | "otro" | null,
  "cobro_cancelacion": null,
  "ambiguo": false,
  "nota": null
}}"""

def parse_email_with_claude(subject, body):
    prompt = PARSE_PROMPT_TEMPLATE.format(subject=subject, body=body[:2500])

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if api_key:
        try:
            import anthropic
            msg = anthropic.Anthropic(api_key=api_key, timeout=60).messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=512,
                messages=[{"role": "user", "content": prompt}],
            )
            text = msg.content[0].text.strip()
            m = re.search(r"\{.*\}", text, re.DOTALL)
            return json.loads(m.group() if m else text)
        except Exception as e:
            return {"tipo": "irrelevante", "nota": f"api_error: {e}"}

    if not CLAUDE_BIN:
        return {"tipo": "irrelevante", "nota": "claude CLI no encontrado y sin ANTHROPIC_API_KEY"}
    try:
        result = subprocess.run(
            [CLAUDE_BIN, "-p", prompt, "--dangerously-skip-permissions"],
            capture_output=True, text=True, timeout=90,
        )
        text = result.stdout.strip()
        m = re.search(r"\{.*\}", text, re.DOTALL)
        return json.loads(m.group() if m else text)
    except Exception as e:
        return {"tipo": "irrelevante", "nota": f"parse_error: {e}"}


# ── Google auth ───────────────────────────────────────────────────────────────

def get_google_creds():
    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDS_FILE), SCOPES)
            creds = flow.run_local_server(port=0)
        TOKEN_FILE.write_text(creds.to_json())
    return creds


# ── Gmail helpers ─────────────────────────────────────────────────────────────

def search_threads(svc, query):
    result = svc.users().threads().list(userId="me", q=query).execute()
    return [t["id"] for t in result.get("threads", [])]

def thread_subject(svc, tid):
    t = svc.users().threads().get(
        userId="me", id=tid, format="metadata", metadataHeaders=["Subject"]
    ).execute()
    for msg in t.get("messages", [])[:1]:
        for h in msg.get("payload", {}).get("headers", []):
            if h["name"] == "Subject":
                return h["value"]
    return ""

def extract_text(payload, max_chars=3000):
    mime = payload.get("mimeType", "")
    data = payload.get("body", {}).get("data", "")
    if data and mime in ("text/plain", "text/html"):
        raw = base64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace")
        if mime == "text/html":
            raw = re.sub(r"<[^>]+>", " ", raw)
        return re.sub(r"\s+", " ", raw).strip()[:max_chars]
    for part in payload.get("parts", []):
        result = extract_text(part, max_chars)
        if result:
            return result
    return ""

def get_thread_text(svc, tid):
    thread = svc.users().threads().get(userId="me", id=tid, format="full").execute()
    return "\n---\n".join(
        t for msg in thread.get("messages", [])
        if (t := extract_text(msg.get("payload", {})))
    )


# ── Excel helpers ─────────────────────────────────────────────────────────────

def parse_date(s):
    if not s:
        return None
    if isinstance(s, datetime.datetime):
        return s
    if isinstance(s, datetime.date):
        return datetime.datetime.combine(s, datetime.time())
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.datetime.strptime(str(s), fmt)
        except ValueError:
            pass
    return None


def _to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None

def is_separator(ws, r):
    v = str(ws.cell(r, 1).value or "") + str(ws.cell(r, 3).value or "")
    return any(x in v.upper() for x in ["TASAS", "BLOQUEO", "VASILE", "TOTAL"])

def find_total_row(ws):
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(r, 1).value or "").strip().upper() == "TOTAL":
            return r
    return ws.max_row

_ROW_FONT   = Font(name="Arial", size=11)
_ROW_BORDER = Border(
    left=Side(style="thin", color="FF000000"),
    right=Side(style="thin", color="FF000000"),
    top=Side(style="thin", color="FF000000"),
    bottom=Side(style="thin", color="FF000000"),
)

def apply_row_style(ws, r, fecha_entrada, today):
    color = GREEN if (fecha_entrada and fecha_entrada.date() <= today) else WHITE
    fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws.row_dimensions[r].height = 18
    for col, fmt, align in COL_FORMATS:
        cell = ws.cell(r, col)
        cell.number_format = fmt
        cell.alignment     = Alignment(horizontal=align)
        cell.fill          = fill
        cell.font          = _ROW_FONT
        cell.border        = _ROW_BORDER

def update_sum_formulas(ws, total_row):
    for letter, col in [("D", 4), ("E", 5), ("I", 9)]:
        ws.cell(total_row, col).value = f"=SUM({letter}2:{letter}{total_row-1})"

def find_guest_row(ws, nombre, fecha_entrada=None):
    if not nombre:
        return None
    nl = nombre.lower()
    for r in range(2, ws.max_row + 1):
        cn = str(ws.cell(r, 3).value or "").lower()
        if nl in cn or cn in nl:
            if fecha_entrada is None:
                return r
            d = ws.cell(r, 1).value
            if d and hasattr(d, "date") and d.date() == fecha_entrada.date():
                return r
    return None

def insert_reservation(ws, data, today):
    fecha_e = parse_date(data.get("fecha_entrada"))
    fecha_s = parse_date(data.get("fecha_salida"))
    if not fecha_e:
        return False
    if find_guest_row(ws, data.get("nombre"), fecha_e):
        return False  # already exists

    total_row = find_total_row(ws)
    insert_at = total_row
    for r in range(2, total_row):
        if is_separator(ws, r):
            continue
        d = ws.cell(r, 1).value
        if d and hasattr(d, "date") and d.date() > fecha_e.date():
            insert_at = r
            break

    ws.insert_rows(insert_at)
    total_row += 1

    # Booking: importe se rellena manualmente al final de la estancia — se deja en blanco
    es_booking = str(data.get("plataforma") or "").strip().lower() == "booking"
    total_val   = None if es_booking else data.get("total_eur")

    for col, val in enumerate([
        fecha_e, fecha_s, data.get("nombre"),
        data.get("noches"), total_val, data.get("plataforma"),
        data.get("adultos") or 0, data.get("ninos") or 0,
        f"=IF(G{insert_at}=0,0,1.75*G{insert_at}*MIN(D{insert_at},7))",
        None, None, None,  # col 10=COBRO TASA, col 11=PASADA A CUENTA, col 12=EXTRAS
    ], 1):
        ws.cell(insert_at, col).value = val

    apply_row_style(ws, insert_at, fecha_e, today)
    update_sum_formulas(ws, total_row)
    return True

def refresh_colors(ws, today):
    for r in range(2, ws.max_row + 1):
        vc = (str(ws.cell(r, 1).value or "") + str(ws.cell(r, 3).value or "")).upper()
        if "TOTAL" in vc and not any(k in vc for k in ("TASAS", "BLOQUEO")):
            fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        elif is_separator(ws, r):
            continue  # TASAS/BLOQUEO — preservar color existente
        else:
            d = ws.cell(r, 1).value
            if not (d and hasattr(d, "date")):
                continue
            color = GREEN if d.date() <= today else WHITE
            fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, ws.max_column + 1):
            ws.cell(r, col).fill = fill


# ── Drive helpers ─────────────────────────────────────────────────────────────

def download_from_drive(svc, local_path, parent_id):
    """Descarga el fichero de Drive y sobreescribe la copia local.
    Devuelve True si se descargó, False si no existe en Drive todavía."""
    name = Path(local_path).name
    existing = svc.files().list(
        q=f"name='{name}' and '{parent_id}' in parents and trashed=false",
        fields="files(id,modifiedTime)", supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute().get("files", [])
    if not existing:
        return False
    content = svc.files().get_media(
        fileId=existing[0]["id"], supportsAllDrives=True
    ).execute()
    Path(local_path).write_bytes(content)
    return True


def upload_file(svc, local_path, parent_id):
    name = Path(local_path).name
    existing = svc.files().list(
        q=f"name='{name}' and '{parent_id}' in parents and trashed=false",
        fields="files(id)", supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute().get("files", [])
    with open(local_path, "rb") as f:
        media = MediaIoBaseUpload(io.BytesIO(f.read()), mimetype=XLSX_MIME, resumable=False)
    if existing:
        svc.files().update(
            fileId=existing[0]["id"], media_body=media, supportsAllDrives=True
        ).execute()
    else:
        svc.files().create(
            body={"name": name, "parents": [parent_id], "mimeType": XLSX_MIME},
            media_body=media, supportsAllDrives=True,
        ).execute()

_CELL_BORDER = Border(
    left=Side(style="thin", color="FF000000"),
    right=Side(style="thin", color="FF000000"),
    top=Side(style="thin", color="FF000000"),
    bottom=Side(style="thin", color="FF000000"),
)

def insert_limp_row(ws, apt_drive_name, checkout_dt):
    """Inserta fila en LIMPIEZAS_PROXIMAS ordenada por fecha con colores y bordes."""
    apt_upper   = apt_drive_name.upper()
    checkout_date = _to_date(checkout_dt)
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == apt_upper and _to_date(ws.cell(r, 2).value) == checkout_date:
            return False

    insert_at = ws.max_row + 1
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 2).value
        if d and hasattr(d, "date") and d > checkout_dt:
            insert_at = r
            break

    ws.insert_rows(insert_at)

    apt_fill  = PatternFill(start_color=APT_COLORS.get(apt_upper, "FFFFFFFF"),
                            end_color=APT_COLORS.get(apt_upper, "FFFFFFFF"), fill_type="solid")
    grey_fill = PatternFill(start_color=LIMP_GREY, end_color=LIMP_GREY, fill_type="solid")
    dia = DIAS_ES[checkout_dt.weekday()] if hasattr(checkout_dt, "weekday") else ""

    vals = [apt_upper, checkout_dt, "", dia, None]
    fills = [apt_fill, grey_fill, grey_fill, grey_fill, grey_fill]
    fmts  = [None, "dd/mm/yyyy", None, None, None]
    bolds = [True, False, False, False, False]

    for i, (val, fill, fmt, bold) in enumerate(zip(vals, fills, fmts, bolds), 1):
        c = ws.cell(insert_at, i)
        if val is not None: c.value = val
        c.fill      = fill
        c.border    = _CELL_BORDER
        c.alignment = Alignment(horizontal="center")
        if fmt:  c.number_format = fmt
        if bold: c.font = Font(bold=True)

    return True


def refresh_limp_colors(ws_l, today):
    """Colorea de gris toda la fila cuando la limpieza ya se realizó (fecha < hoy)."""
    done_fill = PatternFill(start_color=LIMP_HECHA, end_color=LIMP_HECHA, fill_type="solid")
    changed = False
    for r in range(2, ws_l.max_row + 1):
        d = ws_l.cell(r, 2).value
        d_date = _to_date(d)
        if not d_date:
            continue
        if d_date >= today:
            continue
        for col in range(1, 6):
            cell = ws_l.cell(r, col)
            current = cell.fill.fgColor.rgb if cell.fill and cell.fill.fill_type == "solid" else ""
            if current != LIMP_HECHA:
                cell.fill = done_fill
                changed = True
    return changed


def refresh_montaje(ws_l, wb_r, today=None):
    """Actualiza MONTAJE de cada fila con el huésped siguiente; respeta ediciones manuales."""
    if today is None:
        today = datetime.date.today()
    # Construir: apt_upper -> lista de (fecha_entrada, adultos, ninos, extras) ordenada
    apt_entries = {}
    for sheet_name, apt_key in DRIVE_MAP.items():
        if sheet_name not in wb_r.sheetnames:
            continue
        ws = wb_r[sheet_name]
        cols = SHEET_COL_MAP.get(sheet_name, _DEFAULT_COLS)
        rows = []
        for r in range(2, ws.max_row + 1):
            fe = ws.cell(r, 1).value
            if not (fe and hasattr(fe, "date")):
                continue
            try:
                adultos = int(ws.cell(r, cols["adultos"]).value or 0)
                ninos   = int(ws.cell(r, cols["ninos"]).value or 0)
            except (TypeError, ValueError):
                adultos, ninos = 0, 0
            extras = str(ws.cell(r, cols["extras"]).value or "")
            rows.append((fe.date(), adultos, ninos, extras))
        apt_entries[apt_key] = sorted(rows, key=lambda x: x[0])

    changed = False
    for r in range(2, ws_l.max_row + 1):
        apt   = str(ws_l.cell(r, 1).value or "").upper()
        fecha = ws_l.cell(r, 2).value
        if not (apt and fecha and hasattr(fecha, "date")):
            continue
        current = str(ws_l.cell(r, 3).value or "").strip()
        if current not in _AUTO_MONTAJE:
            continue  # edición manual — no tocar
        adultos, ninos, extras = 0, 0, ""
        for entrada, a, n, e in apt_entries.get(apt, []):
            if entrada >= fecha.date():
                adultos, ninos, extras = a, n, e
                break
        nuevo = camas_label(adultos, ninos, extras)
        if nuevo == current:
            continue
        # Nunca bajar de 2 CAMAS a 1 CAMA — solo upgrades
        if nuevo == "1 CAMA" and current == "2 CAMAS":
            continue
        ws_l.cell(r, 3).value = nuevo
        changed = True
    return changed




# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--days", type=int, default=4, help="Días hacia atrás para buscar emails")
    args = parser.parse_args()
    days = args.days

    today = datetime.date.today()
    lines = [
        "\n" + "=" * 60,
        f"Ejecución: {datetime.datetime.now():%Y-%m-%d %H:%M:%S}  (ventana: {days}d)",
        "=" * 60,
    ]

    if not CLAUDE_BIN and not os.environ.get("ANTHROPIC_API_KEY"):
        lines.append("❌ ERROR: se necesita claude CLI o ANTHROPIC_API_KEY.")
        _write_log(lines)
        sys.exit(1)

    if not CREDS_FILE.exists():
        lines.append("❌ ERROR: falta credentials.json. Sigue las instrucciones de setup.")
        _write_log(lines)
        sys.exit(1)

    # Auth Google
    creds = get_google_creds()
    import httplib2
    from google_auth_httplib2 import AuthorizedHttp as _AuthHttp
    _http = _AuthHttp(creds, http=httplib2.Http(timeout=30))
    gmail = build("gmail", "v1", http=_http)
    drive = build("drive", "v3", http=_http)

    # Sincronizar desde Drive antes de tocar nada — los cambios manuales del Drive mandan
    try:
        r_ok = _retry(lambda: download_from_drive(drive, str(RESERVAS_XLS), DRIVE_RESERVAS_PARENT))
    except Exception as exc:
        lines.append(f"  ⚠️  No se pudo descargar RESERVAS de Drive: {exc}. Usando copia local.")
        r_ok = RESERVAS_XLS.exists()
    try:
        l_ok = _retry(lambda: download_from_drive(drive, str(LIMPIEZAS_XLS), DRIVE_LIMPIEZAS_PARENT))
    except Exception as exc:
        lines.append(f"  ⚠️  No se pudo descargar LIMPIEZAS de Drive: {exc}. Usando copia local.")
        l_ok = LIMPIEZAS_XLS.exists()
    lines.append(f"📥 Drive → local: RESERVAS={'✅' if r_ok else '⚠️ no encontrado'}, LIMPIEZAS={'✅' if l_ok else '⚠️ no encontrado'}")

    # Cargar Excel
    wb_r = openpyxl.load_workbook(str(RESERVAS_XLS))
    wb_l = openpyxl.load_workbook(str(LIMPIEZAS_XLS))
    ws_l = wb_l["ACTUAL y próximas"]

    reservas_changed  = False
    limpiezas_changed = False
    nuevas, canceladas, extras_log, ambiguos = [], [], [], []

    # Buscar emails
    queries = [
        f"from:airbnb.com newer_than:{days}d",
        f"from:booking.com newer_than:{days}d",
        f"(from:holidu.com OR from:holidu.de) newer_than:{days}d",
        f"(pago recibido OR payment received OR transferencia) newer_than:{days}d",
    ]
    seen, thread_ids = set(), []
    for q in queries:
        for tid in search_threads(gmail, q):
            if tid not in seen:
                seen.add(tid)
                thread_ids.append(tid)

    # Emails de reservas directas
    direct_thread_ids = set()
    direct_days = max(days, 7)
    for tid in search_threads(gmail, f"from:calafellapartament@gmail.com newer_than:{direct_days}d"):
        direct_thread_ids.add(tid)
        if tid not in seen:
            seen.add(tid)
            thread_ids.append(tid)

    lines.append(f"Threads encontrados: {len(thread_ids)}")

    for tid in thread_ids:
        try:
            subject = _retry(lambda t=tid: thread_subject(gmail, t))
            body    = _retry(lambda t=tid: get_thread_text(gmail, t))
        except Exception as exc:
            lines.append(f"  ⚠️  Thread {tid}: error de red tras 3 intentos ({exc}), ignorado")
            continue
        if not subject.strip() and not body.strip():
            lines.append(f"  ⚠️  Thread {tid}: sin contenido extraible, ignorado")
            continue
        data    = parse_email_with_claude(subject, body)

        nota = data.get("nota", "")
        if nota and ("parse_error" in str(nota) or "api_error" in str(nota)):
            lines.append(f"  ⚠️  Thread {tid} ({subject[:50]}): {nota}")

        # Forzar plataforma "Directa" para confirmaciones propias
        if tid in direct_thread_ids and data.get("tipo") == "nueva_reserva":
            data["plataforma"] = "Directa"

        tipo = data.get("tipo", "irrelevante")
        apto = data.get("apartamento")
        lines.append(f"  📧 {subject[:60]}: {tipo}/{apto or '–'}")
        if tipo == "irrelevante" or not apto:
            continue

        sheet_name = f"{apto} 2026"
        if sheet_name not in wb_r.sheetnames:
            lines.append(f"  ⚠️  Apartamento no reconocido: {apto}")
            continue

        # Open Sky y Duplex se gestionan manualmente en Drive — no procesar emails
        if sheet_name in OS_DUPLEX_SHEETS:
            lines.append(
                f"  ⚠️  Email clasificado como {apto} ({tipo}) — "
                f"este apartamento no recibe reservas por email. "
                f"Revisa el email manualmente: {data.get('nombre')} {data.get('fecha_entrada')}"
            )
            continue

        ws = wb_r[sheet_name]
        if data.get("ambiguo"):
            ambiguos.append(f"{apto}: {data.get('nota', subject[:60])}")

        if tipo == "nueva_reserva":
            if insert_reservation(ws, data, today):
                reservas_changed = True
                nuevas.append(
                    f"{apto} | {data.get('nombre')} | "
                    f"{data.get('fecha_entrada')}–{data.get('fecha_salida')} | "
                    f"{data.get('total_eur')}€"
                )
                fecha_s_dt = parse_date(data.get("fecha_salida"))
                if fecha_s_dt and fecha_s_dt.date() >= today:
                    apt_drive = DRIVE_MAP[sheet_name]
                    if insert_limp_row(ws_l, apt_drive, fecha_s_dt):
                        limpiezas_changed = True

        elif tipo == "cancelacion":
            fecha_e_dt = parse_date(data.get("fecha_entrada"))
            row_num = find_guest_row(ws, data.get("nombre"), fecha_e_dt)
            if row_num and row_num < find_total_row(ws):
                # Reservas directas no se tocan bajo ningún concepto
                if str(ws.cell(row_num, 6).value or "").lower() == "directa":
                    lines.append(f"  ⚠️  Cancelación ignorada (reserva directa): {data.get('nombre')}")
                    continue
                fecha_s_val  = ws.cell(row_num, 2).value
                fecha_s_date = _to_date(fecha_s_val)
                apt_drive = DRIVE_MAP.get(sheet_name, "")
                try:
                    cobro = float(data.get("cobro_cancelacion") or 0)
                except (TypeError, ValueError):
                    cobro = 0

                if cobro > 0:
                    # Cancelación parcial: mantener fila, actualizar importe y marcar
                    ws.cell(row_num, 5).value  = cobro
                    ws.cell(row_num, 9).value  = 0
                    ws.cell(row_num, 12).value = "CANCELADA"
                    reservas_changed = True
                    canceladas.append(f"{apto} | {data.get('nombre')} | {data.get('fecha_entrada')} | cobro {cobro}€")
                else:
                    # Cancelación total: eliminar fila
                    ws.delete_rows(row_num)
                    update_sum_formulas(ws, find_total_row(ws))
                    reservas_changed = True
                    canceladas.append(f"{apto} | {data.get('nombre')} | {data.get('fecha_entrada')}")

                # En ambos casos: eliminar la limpieza de salida del huésped cancelado
                for r in range(ws_l.max_row, 1, -1):
                    if (fecha_s_date
                            and str(ws_l.cell(r, 1).value or "").upper() == apt_drive.upper()
                            and _to_date(ws_l.cell(r, 2).value) == fecha_s_date):
                        ws_l.delete_rows(r)
                        limpiezas_changed = True
                        break

        elif tipo == "modificacion":
            # Buscar la fila por nombre + fecha anterior (o entrada nueva si no hay anterior)
            fecha_ant = parse_date(data.get("fecha_entrada_anterior")) or parse_date(data.get("fecha_entrada"))
            row_num = find_guest_row(ws, data.get("nombre"), fecha_ant)
            if not row_num:
                row_num = find_guest_row(ws, data.get("nombre"))
            if row_num:
                if str(ws.cell(row_num, 6).value or "").lower() == "directa":
                    lines.append(f"  ⚠️  Modificación ignorada (reserva directa): {data.get('nombre')}")
                    continue
                fecha_e_nueva = parse_date(data.get("fecha_entrada"))
                fecha_s_nueva = parse_date(data.get("fecha_salida"))
                vieja_salida  = ws.cell(row_num, 2).value
                es_booking = str(ws.cell(row_num, 6).value or "").strip().lower() == "booking"
                if fecha_e_nueva:
                    ws.cell(row_num, 1).value = datetime.datetime.combine(fecha_e_nueva, datetime.time())
                if fecha_s_nueva:
                    ws.cell(row_num, 2).value = datetime.datetime.combine(fecha_s_nueva, datetime.time())
                if data.get("noches"):
                    ws.cell(row_num, 4).value = data["noches"]
                if data.get("total_eur") and not es_booking:
                    ws.cell(row_num, 5).value = data["total_eur"]
                if data.get("adultos"):
                    ws.cell(row_num, 7).value = data["adultos"]
                if data.get("ninos") is not None:
                    ws.cell(row_num, 8).value = data["ninos"]
                apply_row_style(ws, row_num, fecha_e_nueva or parse_date(ws.cell(row_num, 1).value), today)
                reservas_changed = True
                nuevas.append(
                    f"MODIF {apto} | {data.get('nombre')} | "
                    f"{data.get('fecha_entrada')}–{data.get('fecha_salida')} | "
                    f"{data.get('total_eur')}€"
                )
                # Actualizar limpieza si cambió la fecha de salida
                if fecha_s_nueva and vieja_salida and fecha_s_nueva != (vieja_salida.date() if hasattr(vieja_salida, "date") else vieja_salida):
                    for r in range(ws_l.max_row, 1, -1):
                        if ws_l.cell(r, 2).value == vieja_salida:
                            ws_l.cell(r, 2).value = datetime.datetime.combine(fecha_s_nueva, datetime.time())
                            limpiezas_changed = True
                            break

        elif tipo == "extra":
            row_num = find_guest_row(ws, data.get("nombre"))
            if row_num:
                # Reservas directas no se tocan bajo ningún concepto
                if str(ws.cell(row_num, 6).value or "").lower() == "directa":
                    continue
                canal = data.get("extra_canal") or "airbnb"
                # Extras de PayPal/Revolut los lleva la propietaria manualmente — no tocar
                if canal in ("paypal", "revolut"):
                    continue
                # No sobreescribir si ya hay un valor anotado manualmente
                existing_extra = ws.cell(row_num, 12).value
                if existing_extra:
                    continue
                # Ambiguo sin importe claro → no escribir nada, ya se notificó por Telegram
                importe = data.get("extra_importe")
                if not importe:
                    continue
                ws.cell(row_num, 12).value = f"{importe}€ {canal}"
                extras_log.append(f"{apto} | {data.get('nombre')} | {importe}€ {canal}")
                reservas_changed = True

    # Actualizar colores
    for sname in SHEETS:
        if sname in wb_r.sheetnames:
            refresh_colors(wb_r[sname], today)

    # Colorear en gris las limpiezas ya realizadas (fecha pasada) — nunca se borran
    if refresh_limp_colors(ws_l, today):
        limpiezas_changed = True

    # Recalcular MONTAJE con la reserva siguiente (respeta ediciones manuales)
    if refresh_montaje(ws_l, wb_r, today):
        limpiezas_changed = True

    # Guardar y subir solo si hubo cambios (evita sobreescribir ediciones manuales en Drive)
    if reservas_changed:
        wb_r.save(str(RESERVAS_XLS))
        try:
            _retry(lambda: upload_file(drive, str(RESERVAS_XLS), DRIVE_RESERVAS_PARENT))
            lines.append("✅ RESERVAS_2026.xlsx subido a Drive")
        except Exception as exc:
            lines.append(f"  ⚠️  RESERVAS guardado localmente pero no subido a Drive: {exc}")
    else:
        lines.append("ℹ️  RESERVAS sin cambios — no se sube a Drive")

    if limpiezas_changed:
        wb_l.save(str(LIMPIEZAS_XLS))
        try:
            _retry(lambda: upload_file(drive, str(LIMPIEZAS_XLS), DRIVE_LIMPIEZAS_PARENT))
            lines.append("✅ LIMPIEZAS_2026.xlsx subido a Drive")
        except Exception as exc:
            lines.append(f"  ⚠️  LIMPIEZAS guardado localmente pero no subido a Drive: {exc}")
    else:
        lines.append("ℹ️  LIMPIEZAS sin cambios — no se sube a Drive")

    # Resumen
    if not (nuevas or canceladas or extras_log):
        lines.append("\nSin cambios en las últimas 48h.")
    else:
        lines.append("\n📋 RESUMEN")
        for label, items, icon in [
            ("Reservas nuevas", nuevas,     "✅"),
            ("Cancelaciones",   canceladas, "❌"),
            ("Extras",          extras_log, "💶"),
        ]:
            lines.append(f"  {label}: {len(items)}")
            for item in items:
                lines.append(f"    {icon} {item}")
    if ambiguos:
        lines.append("  ⚠️ Ambiguos para revisión manual:")
        for a in ambiguos:
            lines.append(f"    {a}")
        msg = "⚠️ <b>Pago sin clasificar — revisión manual</b>\n\n"
        msg += "\n".join(f"• {a}" for a in ambiguos)
        msg += "\n\n¿Es lavandería (25€), mascota (100€), check-in anticipado (30€) u otro concepto? Apúntalo manualmente en el Excel."
        send_telegram(msg)

    lines.append("=" * 60 + "\n")
    _write_log(lines)

    # Exportar caché JSON para el bot de Telegram y sincronizar a Drive
    _write_bot_cache(wb_r, wb_l, drive)


def _write_bot_cache(wb_r, wb_l, drive=None):
    """Exporta reservas y limpiezas a JSON accesible por el bot, y lo sube a Drive."""
    today = datetime.date.today()

    def _num(v):
        try: return float(v or 0)
        except (TypeError, ValueError): return 0.0

    reservas = []
    for sname in SHEETS:
        if sname not in wb_r.sheetnames:
            continue
        apt  = sname.replace(" 2026", "")
        ws   = wb_r[sname]
        cols = SHEET_COL_MAP.get(sname, _DEFAULT_COLS)
        is_os_dup = sname in OS_DUPLEX_SHEETS
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre = str(row[2] or "").strip()
            skip_words = ["TOTAL", "TASAS", "BLOQUEO", "BLOQ"]
            if not nombre or any(x in nombre.upper() for x in skip_words):
                continue
            fe = row[0]
            fs = row[1]
            if not fe or not hasattr(fe, "strftime"):
                continue
            if is_os_dup:
                # Estructura OS/Duplex: col5=ADULTOS, col6=NINOS, col8=PLATAFORMA, col9=TOTAL, col10=EXTRAS
                reservas.append({
                    "apt":      apt,
                    "nombre":   nombre,
                    "entrada":  fe.strftime("%Y-%m-%d"),
                    "salida":   fs.strftime("%Y-%m-%d") if (fs and hasattr(fs, "strftime")) else str(fs or ""),
                    "noches":   int(_num(row[3])),
                    "total":    _num(row[8]),   # TOTAL (col 9, idx 8)
                    "platform": str(row[7] or "").strip() or "—",  # PLATAFORMA (col 8, idx 7)
                    "adultos":  int(_num(row[cols["adultos"]-1])),
                    "ninos":    int(_num(row[cols["ninos"]-1])),
                    "tasas":    0.0,
                    "extras":   str(row[cols["extras"]-1] or "").strip(),  # EXTRAS (col 10, idx 9)
                })
            else:
                reservas.append({
                    "apt":      apt,
                    "nombre":   nombre,
                    "entrada":  fe.strftime("%Y-%m-%d"),
                    "salida":   fs.strftime("%Y-%m-%d") if (fs and hasattr(fs, "strftime")) else str(fs or ""),
                    "noches":   int(_num(row[3])),
                    "total":    _num(row[4]),
                    "platform": str(row[5] or "").strip() or "—",
                    "adultos":  int(_num(row[6])),
                    "ninos":    int(_num(row[7])),
                    "tasas":    _num(row[8]),
                    "extras":   str(row[11] or "").strip(),
                })

    limpiezas = []
    try:
        ws_l = wb_l["ACTUAL y próximas"]
        for row in ws_l.iter_rows(min_row=2, values_only=True):
            apt   = str(row[0] or "").strip()
            fecha = row[1]
            if not apt or not fecha:
                continue
            fecha_date = fecha.date() if hasattr(fecha, "date") else None
            limpiezas.append({
                "apt":     apt,
                "fecha":   fecha_date.strftime("%Y-%m-%d") if fecha_date else str(fecha),
                "montaje": str(row[2] or "").strip(),
                "dia":     str(row[3] or "").strip(),
                "hecha":   bool(fecha_date and fecha_date < today),
            })
    except Exception:
        pass

    cache = {
        "updated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "reservas": reservas,
        "limpiezas": limpiezas,
    }
    cache_path = SCRIPT_DIR / "data_cache.json"
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    if drive:
        try:
            upload_file(drive, str(cache_path), DRIVE_CACHE_PARENT)
        except Exception:
            pass


def _write_log(lines):
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(str(LOG_FILE), "a", encoding="utf-8") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    main()
